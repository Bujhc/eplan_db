from openpyxl import Workbook, load_workbook

import pyodbc

def request_db(product_db, price):
    for i in price:
        if i[0] == product_db[0]:
            print(i)
            print('price for {0} is {1}'.format (product_db[0], i[1] ))
    

workbook = load_workbook(filename="CSPEC Oleg.xlsx")
sheet = workbook.active
products= []


for table_line in sheet.iter_rows(min_row=1, min_col=1, max_col=5, values_only=True):
    convert_line = list(table_line[0:5:4])
    products.append(convert_line)
    #print(type(products))

product_1 = products
#print(product_1[0][0])

conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)}; DBQ=C:\Project\eplan\parts.mdb;')
cursor = conn.cursor()

conn.execute("update tblPart set salesprice1='33333' where partnr='MS-NCE2511-0'")
conn.commit()

cursor.execute("""
                select partnr, lastchange from tblPart
                where manufacturer='Johnson Controls'
                """)

row = cursor.fetchall()
print(row)
#row_1 = row[50]
print(set(product_1).intersection(row))
#request_db(row, product_1)





conn.close