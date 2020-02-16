import glob
import os
import pyodbc

from openpyxl import Workbook, load_workbook
from pathlib import Path


# find items in price table which have in db of eplan
def request_db(product_db, price):
    
    finish_price= []
    
    for i in price:                                # prepare list with price for SQL request/update
        for y in product_db:
                if i[0] == y[0]:
                    index = price.index(i)
                    finish_price.append(price[index])
    
    return finish_price

def update_db(list_with_price):
    
    for items in list_with_price:
        partnr = repr(str(items[0]))
        salesprice1 = repr(str(items[1]))
        sql_update_request = "update tblPart set salesprice1=? where partnr=?"
        cursor.execute(sql_update_request,salesprice1,partnr)
        conn.commit
        


# condition ZERO

BRAND_LIST = []



# first request to define how many brands have in thr database
# the path where locate database
    
conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)}; DBQ=C:\Users\Bujhc\eplan_db\parts.mdb;')
cursor = conn.cursor()

sql=("""
                select distinct manufacturer from tblPart where manufacturer IS NOT NULL and manufacturer <> ' ' 
                
                """)

cursor.execute(sql)

for i in cursor.fetchall():
    
    BRAND_LIST.append(list(i))

conn.close

# the path where locate price list

os.chdir(r"D:\Dropbox\Прайс-лист")

path_for_brand = r"D:\Dropbox\Прайс-лист"

spisok_folder_brand_price = []
dict_of_path_price = {}

# we prepare dictionary of pathes with key like name of brand
# it is easy to use in the next part of code, for example , we call name of brand and if you know the path where locate price list
# you can use another library for save items of price in the database

for name_folder_brand in os.listdir(path_for_brand):
    path_request = os.path.join(path_for_brand, name_folder_brand)
    spisok_folder_brand_price.append(name_folder_brand)
    dict_of_path_price.update({name_folder_brand : path_request})
   

for filename in Path().rglob('*.xl*x'):

    filepath = os.path.abspath(filename)
    #print(filepath)

finish_price = []

# in this part of code we start the main loop part of programm
for i in BRAND_LIST:
    for NAME_BRANCH in spisok_folder_brand_price:
        if i[0] == NAME_BRANCH:
            print(dict_of_path_price[NAME_BRANCH])
            
            # разобраться как ставить путь + привести все таблицы в порядок
            workbook = load_workbook(filename=r"C:\Users\Bujhc\eplan_db\CSPEC_Oleg.xlsx", read_only=True)
            sheet = workbook.active
            price= []

            for table_line in sheet.iter_rows(min_row=1, min_col=1, max_col=5, values_only=True):
                convert_line = list(table_line[0:5:4])
                price.append(convert_line)
            try:    
                
                conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=C:\Users\Bujhc\eplan_db\parts.mdb;')
                conn.autocommit = True                      # very important flag for db - allow to write in access 
                cursor = conn.cursor()


                sql=("""
                    SELECT partnr, salesprice1 
                    FROM tblPart
                    WHERE manufacturer=?
                    """)

                row = cursor.execute(sql,NAME_BRANCH)

                update_db(request_db(row, price))

                sql1=("""
                    SELECT partnr, salesprice1, purchaseprice1
                    FROM tblPart
                    WHERE manufacturer=?
                    """)

                cursor.execute(sql1,NAME_BRANCH)

                row1 = cursor.fetchall()
                for i in row1:
                    print(i)

            except pyodbc.Error as e:
                loggin.error(e)
                print(e)

            cursor.close()
            conn.close()