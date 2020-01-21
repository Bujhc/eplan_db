import openpyxl
import pandas as pd
import numpy as np

from itertools import islice
from openpyxl import Workbook, load_workbook, cell
from openpyxl.utils import column_index_from_string, column_index_from_string, get_column_letter

workbook = load_workbook(filename="BOQ.xlsx")

sheet = workbook.active

df = pd.DataFrame(sheet.values)

data = sheet.values

cols = next(data)[1:]
data = list(data)
idx = [r[0] for r in data]

data = (islice(r,1,None) for r in data)

df = pd.DataFrame(data, index=idx, columns=cols)

df_drop_missing = df.dropna()

df_fill = df.fillna(0)

print(df_fill)

















"""
number_cell = 'D40'
#print(sheet[number_cell].value)
#print(sheet['I40'].value)

#sheet['I40'] = 500
product =[]
workbook.save(filename="BOQ.xlsx")

for row in sheet.iter_rows(max_row=20, values_only=True):
    print(row)
    product.append(row[3])
    name_of_material = product[3:4]
    price_material = product[8:9]
    price_of_work = product[9:10]
    

print(product)
print(name_of_material, price_material, price_of_work, )
"""