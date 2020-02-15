import json
from openpyxl import Workbook, load_workbook
import pyodbc
pyodbc.pooling = False

#workbook = load_workbook(filename="CSPEC Oleg.xlsx")

#sheet = workbook.active

#products= {}

#for row in sheet.iter_rows(min_row=2,min_col=1, max_col=5, values_only=True):
 #   if row[0] == "MS-NCE2516-0":
 #       print(row[4])
    
conn = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)}; DBQ=C:\Users\Bujhc\eplan_db\parts.mdb;')
cursor = conn.cursor()

t1 = [['A99DY-200C','111111'],['MS-NAE3510-2', '222222']]


"""
sql_update_request ="update tblPart set salesprice1=? where partnr=?"
conn.execute(sql_update_request,t1[0][1],t1[0][0])
conn.execute(sql_update_request,t1[1][1],t1[1][0])
conn.commit
"""
param = 'Johnson Controls'
sql=("""
                select partnr, purchaseprice1, salesprice1, lastchange from tblPart
                where manufacturer=?
                """)

cursor.execute(sql,param)
    
row = cursor.fetchall()
for i in row:
    print(i)

conn.close